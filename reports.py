import os
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font, Alignment 
from db import get_financial_data
from calculations import calculate_coefficients

def export_to_excel(enterprise_id: int, enterprise_name: str):
    rows = get_financial_data(enterprise_id)
    if not rows:
        messagebox.showerror("Ошибка", "Нет данных для экспорта.")
        return

    # 📁 Папка для сохранения
    reports_folder = os.path.join(os.getcwd(), "reports")
    os.makedirs(reports_folder, exist_ok=True)

    # 📝 Имя файла
    safe_name = enterprise_name.replace(' ', '_')
    filename = os.path.join(reports_folder, f"{safe_name}_report.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Финансовый отчёт"

    headers = [
        "Период",
        "Тек. ликвид.",
        "Автономия",
        "Зависимость",
        "Рентаб. продаж (%)",
        "Рентаб. активов (%)",
        "Оценка"
    ]

    # Заголовки
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for r, row in enumerate(rows, start=2):
        res = calculate_coefficients(row)
        values = [
            res["period"],
            round(res["current_ratio"], 2),
            round(res["autonomy_ratio"], 2),
            round(res["debt_ratio"], 2),
            round(res["profitability_sales"], 2),
            round(res["profitability_assets"], 2),
            res["summary"]
        ]
        for c, val in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=val)

    # Подгон ширины столбцов
    for i in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)
    messagebox.showinfo("Успех", f"Отчёт успешно сохранён в:\n{filename}")
